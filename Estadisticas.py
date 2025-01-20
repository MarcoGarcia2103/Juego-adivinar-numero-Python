#!/usr/bin/env python
# coding: utf-8

# In[1]:


import openpyxl
from datetime import datetime

def guardar_resultados(nombre, resultado, intentos, puntos, nivel_dificultad):
    archivo_excel = "Estadistica.xlsx"
    
    # Cargar o crear el archivo Excel
    try:
        wb = openpyxl.load_workbook(archivo_excel)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
    
    hoja = wb.active
    if hoja.title == "Sheet":  # Cambiar nombre de la hoja si es la predeterminada
        hoja.title = "Resultados"
    
    fila_vacia = hoja.max_row + 1
    fecha_juego = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Escribir los datos en la hoja
    hoja.cell(row=fila_vacia, column=1, value=nombre)  # Columna A: nombre
    hoja.cell(row=fila_vacia, column=2, value=resultado)  # Columna B: resultado
    hoja.cell(row=fila_vacia, column=3, value=intentos)  # Columna C: intentos
    hoja.cell(row=fila_vacia, column=4, value=puntos)  # Columna D: puntos
    hoja.cell(row=fila_vacia, column=5, value=nivel_dificultad)  # Columna E: nivel de dificultad
    hoja.cell(row=fila_vacia, column=6, value=fecha_juego)  # Columna F: fecha
    
    wb.save(archivo_excel)
    wb.close()


# In[2]:




